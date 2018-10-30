#!/home/shiyanlou/anaconda3/bin/python
# -*- coding: utf-8 -*-


from openpyxl import load_workbook # 可以用来载入已有数据表格
from openpyxl import Workbook # 可以用来处理新的数据表格
from openpyxl import worksheet
import datetime # 可以用来处理时间相关的数据
import os

def combine():
    '''
    该函数可以用来处理原数据文件：
    1. 合并表格并写入的 combine 表中
    2. 保存原数据文件
    '''
    wb = load_workbook('courses.xlsx')
    ws1 = wb['students']
    ws2 = wb['time']
    ws3 = wb.create_sheet(title='combine')

    for x, y, z in ws1:
        for a,b,c in ws2:
            if x.value == a.value and y.value == b.value:
                ws3.append((x.value, y.value, z.value, c.value))
    
    wb.save('courses.xlsx') 

def split():

    '''
    该函数可以用来分割文件：
    1. 读取 combine 表中的数据
    2. 将数据按时间分割
    3. 写入不同的数据表中
    '''
    wb = load_workbook('courses.xlsx')
    ws = wb['combine']
    for line in ws:
        if isinstance(line[0].value, str):
            firstline = tuple((i.value for i in line))

        if isinstance(line[0].value, datetime.datetime):  # != '创建时间':
            filename = os.path.abspath(str(line[0].value.year) + '.xlsx')
            if os.path.isfile(filename):
                s_wb = load_workbook(filename)
                s_sb = s_wb.active
            else:
                s_wb = Workbook()
                s_sb = s_wb.active
                s_sb.append(firstline)
                s_wb.save(filename)

            s_sb.append((i.value for i in line))
            s_wb.save(filename)
                

# 执行
if __name__ == '__main__':
    combine()
    split()
